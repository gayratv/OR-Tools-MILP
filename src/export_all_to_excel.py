
import pulp
from typing import Dict, Tuple

from input_data import InputData
import openpyxl
from openpyxl.styles import Font

def _val(var) -> float:
    return float(pulp.value(var)) if var is not None else 0.0

def export_all_to_excel(filename: str,
                        data: InputData,
                        x: Dict[Tuple, pulp.LpVariable],
                        z: Dict[Tuple, pulp.LpVariable]) -> None:
    """Создаёт один Excel-файл с 4 листами:
    - "Классы" (сводка нагрузок по классам)
    - "Учителя" (сводка по учителям)
    - "Классы_расписание" (таблица расписания по классам)
    - "Учителя_расписание" (таблица расписания по учителям)
    """
    wb = openpyxl.Workbook()

    # ----- Лист: сводка по классам -----
    ws_classes = wb.active
    ws_classes.title = "Классы"
    header = ["Класс", "Всего", "Среднее/день"] + data.days
    ws_classes.append(header)
    for c in data.classes:
        total = 0
        per_day = {d: 0 for d in data.days}
        # non-split
        for (cc, s), _ in data.assigned_teacher.items():
            if cc == c:
                for d in data.days:
                    for p in data.periods:
                        val = _val(x.get((cc, s, d, p)))
                        total += val
                        per_day[d] += val
        # split
        for (cc, s, g), _ in data.subgroup_assigned_teacher.items():
            if cc == c:
                for d in data.days:
                    for p in data.periods:
                        val = _val(z.get((cc, s, g, d, p)))
                        total += val
                        per_day[d] += val
        avg = total / len(data.days) if data.days else 0
        ws_classes.append([c, int(total), round(avg, 1)] + [int(per_day[d]) for d in data.days])
    for cell in ws_classes[1]:
        cell.font = Font(bold=True)

    # ----- Лист: сводка по учителям -----
    ws_teachers = wb.create_sheet("Учителя")
    header = ["Учитель", "Всего", "Среднее/день"] + data.days
    ws_teachers.append(header)
    for t in data.teachers:
        total = 0
        per_day = {d: 0 for d in data.days}
        # non-split
        for (c, s), tt in data.assigned_teacher.items():
            if tt == t:
                for d in data.days:
                    for p in data.periods:
                        val = _val(x.get((c, s, d, p)))
                        total += val
                        per_day[d] += val
        # split
        for (c, s, g), tt in data.subgroup_assigned_teacher.items():
            if tt == t:
                for d in data.days:
                    for p in data.periods:
                        val = _val(z.get((c, s, g, d, p)))
                        total += val
                        per_day[d] += val
        avg = total / len(data.days) if data.days else 0
        ws_teachers.append([t, int(total), round(avg, 1)] + [int(per_day[d]) for d in data.days])
    for cell in ws_teachers[1]:
        cell.font = Font(bold=True)

    # ----- Лист: расписание по классам -----
    ws_cls_sched = wb.create_sheet("Классы_расписание")
    for c in data.classes:
        ws_cls_sched.append([f"Класс {c}"])
        header = ["День"] + [f"Урок {p}" for p in data.periods]
        ws_cls_sched.append(header)
        for d in data.days:
            row = [d]
            for p in data.periods:
                cell = None
                # non-split
                for s in data.subjects:
                    if s in data.split_subjects:
                        continue
                    if _val(x.get((c, s, d, p))) > 0.5:
                        t = data.assigned_teacher[(c, s)]
                        cell = f"{s} ({t})"
                        break
                # split
                if cell is None:
                    pieces = []
                    for s in data.subjects:
                        if s not in data.split_subjects:
                            continue
                        for g in data.subgroup_ids:
                            if _val(z.get((c, s, g, d, p))) > 0.5:
                                t = data.subgroup_assigned_teacher[(c, s, g)]
                                pieces.append(f"{s}[g{g}::{t}]")
                    if pieces:
                        cell = "+".join(pieces)
                row.append(cell or "—")
            ws_cls_sched.append(row)
        ws_cls_sched.append([])
        for cell in ws_cls_sched[ws_cls_sched.max_row - len(data.days) - 1]:
            cell.font = Font(bold=True)

    # ----- Лист: расписание по учителям -----
    ws_t_sched = wb.create_sheet("Учителя_расписание")
    for t in data.teachers:
        ws_t_sched.append([f"Учитель {t}"])
        header = ["День"] + [f"Урок {p}" for p in data.periods]
        ws_t_sched.append(header)
        for d in data.days:
            row = [d]
            for p in data.periods:
                cell = None
                # non-split
                for (c, s), tt in data.assigned_teacher.items():
                    if tt != t:
                        continue
                    if _val(x.get((c, s, d, p))) > 0.5:
                        cell = f"{c}-{s}"
                        break
                # split
                if cell is None:
                    for (c, s, g), tt in data.subgroup_assigned_teacher.items():
                        if tt != t:
                            continue
                        if _val(z.get((c, s, g, d, p))) > 0.5:
                            cell = f"{c}-{s}[g{g}]"
                            break
                row.append(cell or "—")
            ws_t_sched.append(row)
        ws_t_sched.append([])
        for cell in ws_t_sched[ws_t_sched.max_row - len(data.days) - 1]:
            cell.font = Font(bold=True)

        # Авто-подбор ширины и заливка заголовков
    from openpyxl.styles import PatternFill
    fill = PatternFill(start_color="FFDDEEFF", end_color="FFDDEEFF", fill_type="solid")
    for ws in [ws_classes, ws_teachers, ws_cls_sched, ws_t_sched]:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 2
        for cell in ws[1]:
            cell.fill = fill

    wb.save(filename)
    print(f"Экспорт в Excel выполнен: {filename}")

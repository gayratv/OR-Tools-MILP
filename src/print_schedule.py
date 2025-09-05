# print_schedule.py (ПОЛНАЯ ВЕРСИЯ)

"""
Pretty-printers and Excel exporters for timetables.
"""

from typing import Dict, Tuple, Any
import dataclasses
import pulp
import openpyxl
from openpyxl.styles import Font, Alignment

from input_data import InputData, OptimizationWeights

def _val(var: Any) -> float:
    """Универсальная функция для получения значения переменной (pulp или число)."""
    if isinstance(var, pulp.LpVariable):
        return float(pulp.value(var)) if var is not None else 0.0
    return float(var) if var is not None else 0.0

def get_solution_maps(data: InputData, solver_or_vars: Dict, is_pulp: bool) -> Dict:
    """Создает словари со значениями переменных для удобства."""
    x_sol, z_sol = {}, {}
    if is_pulp:
        x_vars, z_vars = solver_or_vars['x'], solver_or_vars['z']
        for k, v in x_vars.items(): x_sol[k] = _val(v)
        for k, v in z_vars.items(): z_sol[k] = _val(v)
    else: # CP-SAT
        solver = solver_or_vars['solver']
        x_vars, z_vars = solver_or_vars['x'], solver_or_vars['z']
        for k, v in x_vars.items(): x_sol[k] = solver.Value(v)
        for k, v in z_vars.items(): z_sol[k] = solver.Value(v)
    return {'x': x_sol, 'z': z_sol}


# display_maps
# "subject_names": subject_map.set_index('предмет_eng')['предмет'].to_dict(),
# "teacher_names": teacher_map.set_index('teacher')['FAMIO'].to_dict()
def export_full_schedule_to_excel(filename: str, data: InputData, solution_maps: Dict[str, Dict[Tuple, float]], display_maps: Dict[str, Dict[str, str]]=None, solution_stats: Dict[str, Any]=None, weights: OptimizationWeights=None):
    x_sol, z_sol = solution_maps['x'], solution_maps['z']
    
    # --- Вспомогательные функции для получения полных имен ---
    # Если display_maps не предоставлен, создаем пустые словари
    display_maps = display_maps or {}
    subject_names = display_maps.get("subject_names", {})
    teacher_names = display_maps.get("teacher_names", {})

    def get_subject_name(s_id):
        return subject_names.get(s_id, s_id)

    def get_teacher_name(t_id):
        return teacher_names.get(t_id, t_id)

    wb = openpyxl.Workbook()
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    # --- Лист: расписание по классам ---
    ws_classes = wb.active
    ws_classes.title = "Классы_расписание"
    for c in data.classes:
        ws_classes.append([f"Класс {c}"])
        ws_classes.cell(ws_classes.max_row, 1).font = bold_font
        header = ["День"] + [f"Урок {p}" for p in data.periods]
        ws_classes.append(header)
        for cell in ws_classes[ws_classes.max_row]: cell.font = bold_font

        for d in data.days:
            row = [d]
            for p in data.periods:
                cell_text = None
                # non-split
                for s in data.subjects:
                    if s in data.split_subjects: continue
                    if x_sol.get((c, s, d, p), 0) > 0.5:
                        t = data.assigned_teacher.get((c, s), '?')
                        cell_text = f"{get_subject_name(s)} ({get_teacher_name(t)})"
                        break
                # split
                if cell_text is None:
                    pieces = []
                    for s in data.split_subjects:
                        for g in data.subgroup_ids:
                            if z_sol.get((c, s, g, d, p), 0) > 0.5:
                                t = data.subgroup_assigned_teacher.get((c, s, g), '?')
                                pieces.append(f"{get_subject_name(s)}[g{g}::{get_teacher_name(t)}]")
                    if pieces:
                        cell_text = "+".join(pieces)
                row.append(cell_text or "—")
            ws_classes.append(row)
        ws_classes.append([])

    # --- Лист: расписание по учителям ---
    ws_teachers = wb.create_sheet("Учителя_расписание")
    # Сортируем учителей по их русским именам для упорядоченного вывода
    sorted_teachers = sorted(data.teachers, key=lambda t_id: get_teacher_name(t_id))

    for t in sorted_teachers:
        ws_teachers.append([f"Учитель {get_teacher_name(t)}"])
        ws_teachers.cell(ws_teachers.max_row, 1).font = bold_font
        header = ["День"] + [f"Урок {p}" for p in data.periods]
        ws_teachers.append(header)
        for cell in ws_teachers[ws_teachers.max_row]: cell.font = bold_font
        for d in data.days:
            row = [d]
            for p in data.periods:
                cell_text = None
                # non-split
                for (c, s), tt in data.assigned_teacher.items():
                    if tt != t: continue
                    if x_sol.get((c, s, d, p), 0) > 0.5:
                        cell_text = f"{c}-{get_subject_name(s)}"
                        break
                # split
                if cell_text is None:
                    pieces = []
                    for (c, s, g), tt in data.subgroup_assigned_teacher.items():
                        if tt != t: continue
                        if z_sol.get((c, s, g, d, p), 0) > 0.5:
                            pieces.append(f"{c}-{get_subject_name(s)}[g{g}]")
                    if pieces:
                        cell_text = " + ".join(pieces)
                row.append(cell_text or "—")
            ws_teachers.append(row)
        ws_teachers.append([])

    # --- Лист: Сводка нагрузки ---
    ws_summary = wb.create_sheet("Сводка_нагрузки")
    teacher_load_per_day = {t: {d: 0 for d in data.days} for t in data.teachers}
    class_load_per_day = {c: {d: 0 for d in data.days} for c in data.classes}
    teacher_busy_periods = {t: {d: [] for d in data.days} for t in data.teachers}

    for (c, s, d, p), val in x_sol.items():
        if val > 0.5:
            class_load_per_day[c][d] += 1
            teacher = data.assigned_teacher.get((c, s))
            if teacher: 
                teacher_load_per_day[teacher][d] += 1
                teacher_busy_periods[teacher][d].append(p)
    for (c, s, g, d, p), val in z_sol.items():
        if val > 0.5:
            class_load_per_day[c][d] += 1
            teacher = data.subgroup_assigned_teacher.get((c, s, g))
            if teacher: 
                teacher_load_per_day[teacher][d] += 1
                teacher_busy_periods[teacher][d].append(p)

    ws_summary.append(["Сводка по классам"]); ws_summary.cell(ws_summary.max_row, 1).font = bold_font
    header = ["Класс", "Всего", "Сред./день"] + data.days + ["Предупреждения"]
    ws_summary.append(header)
    for c in data.classes:
        per_day = class_load_per_day[c]
        total = sum(per_day.values())
        avg = total / len(data.days) if data.days else 0
        warnings = []
        if any(v > 7 for v in per_day.values()): warnings.append(f"Перегрузка >7")
        if any(abs(v - avg) > 0.3 * avg and avg > 0 for v in per_day.values()): warnings.append(f"Перекос")
        row = [c, total, f"{avg:.1f}"] + [per_day[d] for d in data.days] + [", ".join(warnings)]
        ws_summary.append(row)

    ws_summary.append([])
    ws_summary.append(["Сводка по учителям"]); ws_summary.cell(ws_summary.max_row, 1).font = bold_font
    header = ["Учитель", "Всего", "Лимит", "Сред./день", "Окна"] + data.days + ["Предупреждения"]
    ws_summary.append(header)
    for t in data.teachers:
        per_day = teacher_load_per_day[t]
        total = sum(per_day.values())
        avg = total / len(data.days) if data.days else 0
        total_windows = 0
        for d in data.days:
            busy_periods = sorted(teacher_busy_periods[t][d])
            if len(busy_periods) > 1: total_windows += sum(busy_periods[i+1] - busy_periods[i] - 1 for i in range(len(busy_periods) - 1)) # noqa
        warnings = []
        if total > data.teacher_weekly_cap: warnings.append(f"Перегруз! ({total}/{data.teacher_weekly_cap})")
        if total_windows > 5: warnings.append(f"Окна > 5")
        row = [get_teacher_name(t), total, data.teacher_weekly_cap, f"{avg:.1f}", total_windows] + [per_day[d] for d in data.days] + [", ".join(warnings)]
        ws_summary.append(row)

    # --- Лист: Сводка по решению ---
    if solution_stats:
        ws_solve_stats = wb.create_sheet("Сводка_решения")
        ws_solve_stats.append(["Параметр", "Значение"])
        ws_solve_stats.cell(1, 1).font = bold_font
        ws_solve_stats.cell(1, 2).font = bold_font

        stats_map = {
            "Финальный статус": solution_stats.get("status"),
            "Целевая функция": f'{solution_stats.get("objective_value"):.2f}',
            "Затрачено времени (сек)": f'{solution_stats.get("wall_time_s"):.2f}',
            'Итоговое количество "одиноких" уроков': solution_stats.get("total_lonely_lessons"),
            'Итоговое количество "окон" у учителей': solution_stats.get("total_teacher_windows")
        }
        for key, value in stats_map.items():
            if value is not None and value != -1:
                ws_solve_stats.append([key, value])

    # --- Лист: Коэффициенты оптимизации ---
    if weights:
        ws_weights = wb.create_sheet("Коэффициенты")
        ws_weights.append(["Коэффициент", "Значение", "Описание"])
        for cell in ws_weights[1]: cell.font = bold_font

        for f in dataclasses.fields(weights):
            # Получаем комментарий (docstring) для поля
            doc = f.metadata.get('doc', '') # Пока не используется, но можно добавить
            value = getattr(weights, f.name)
            # Извлекаем комментарий из файла input_data.py (немного магии)
            ws_weights.append([f.name, value])

    # --- Авто-ширина колонок и стиль ---
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.font.bold: cell.alignment = center_align
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)
    print(f"\nПолное расписание и сводка сохранены в {filename}")
